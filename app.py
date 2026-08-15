import os
import io
import json
import re
import base64
from datetime import datetime, timedelta
import requests
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from PIL import Image, ImageOps
import streamlit as st
from streamlit_cropper import st_cropper
from supabase import create_client, Client
from pypdf import PdfReader, PdfWriter
from pdf2image import convert_from_bytes
import pytesseract
import time

# ReportLab imports per PDF
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# --- CONFIGURAZIONE PAGINA STREAMLIT ---
st.set_page_config(
    page_title="Note Spese 2026",
    page_icon="🧾",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# --- SYSTEM LOGIN ---
def check_password():
    if st.session_state.get("password_correct", False):
        return True

    pwd_input = st.text_input("🔑 Inserisci la Password di accesso", type="password")
    
    if pwd_input:
        if pwd_input == st.secrets.get("PASSWORD"):
            st.session_state["password_correct"] = True
            st.rerun()
        else:
            st.error("😕 Password errata")
            return False
            
    return False

if not check_password():
    st.stop()

# --- CONNESSIONE SUPABASE ---
@st.cache_resource
def init_supabase() -> Client:
    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_KEY"]
    return create_client(url, key)

supabase = init_supabase()

# --- COSTANTI ---
BUCKET_NAME = "allegati-spese"
EXCEL_TEMPLATE = "Note spese 2026_Ferrari.xlsx"
MAX_FILE_SIZE_MB = 5

MANDI_NOMI = [
    "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
    "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"
]

CATEGORIE_LAVORO = [
    "RISTORANTI_CC", "RISTORANTI_CONTANTI",
    "PARCHEGGI_CC", "PARCHEGGI_CONTANTI",
    "CARBURANTE_CC", "CARBURANTE_CARTA",
    "TELEPASS", "NOLO", "ALTRO"
]

CATEGORIE_PERSONALI = [
    "Cibo/Delivery", "Spesa", "Casa", "Auto", 
    "B.", "Romeo", "Personale", "Vacanze", "Varie"
]

CATEGORIA_COLONNA = {
    "TELEPASS": 6,               # Colonna F
    "NOLO": 7,                   # Colonna G
    "PARCHEGGI_CONTANTI": 8,     # Colonna H
    "PARCHEGGI_CC": 9,           # Colonna I
    "RISTORANTI_CONTANTI": 10,   # Colonna J
    "RISTORANTI_CC": 11,         # Colonna K
    "CARBURANTE_CC": 12,         # Colonna L
    "CARBURANTE_CARTA": 13,      # Colonna M
    "ALTRO": 14                  # Colonna N
}

PAGAMENTI_LAVORO = ["CC (Carta)", "Contanti", "Carta Carburante"]
PAGAMENTI_PERSONALI = ["Cash", "Card", "PayPal/Satispay"]

FILL_ORANGE = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# --- UTILITY FOTO OTTIMIZZATE ---

def optimize_pil_image(pil_img, max_dimension=1200, quality=80):
    try:
        pil_img = ImageOps.exif_transpose(pil_img)
        pil_img = pil_img.convert("RGB")

        w, h = pil_img.size
        if max(w, h) > max_dimension:
            ratio = max_dimension / float(max(w, h))
            new_size = (int(w * ratio), int(h * ratio))
            pil_img = pil_img.resize(new_size, Image.Resampling.LANCZOS)

        out_buffer = io.BytesIO()
        pil_img.save(out_buffer, format="JPEG", quality=quality, optimize=True)
        return out_buffer.getvalue()
    except Exception:
        out_buffer = io.BytesIO()
        pil_img.save(out_buffer, format="JPEG")
        return out_buffer.getvalue()

def upload_file_to_supabase(file_bytes, data_spesa, extension, content_type, user_id="default_user"):
    try:
        dt = datetime.strptime(data_spesa, "%Y-%m-%d")
        anno_mese = dt.strftime("%Y-%m")
        now_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        filename = f"rec_{data_spesa}_{now_str}.{extension}"
        storage_path = f"{user_id}/{anno_mese}/{filename}"

        supabase.storage.from_(BUCKET_NAME).upload(
            path=storage_path,
            file=file_bytes,
            file_options={"content-type": content_type}
        )

        return supabase.storage.from_(BUCKET_NAME).get_public_url(storage_path)
    except Exception as e:
        st.error(f"Errore durante l'upload su Supabase Storage: {e}")
        return None

def delete_photo_from_storage(public_url):
    if not public_url or not isinstance(public_url, str):
        return
    try:
        target_token = f"{BUCKET_NAME}/"
        if target_token in public_url:
            storage_path = public_url.split(target_token)[-1]
            supabase.storage.from_(BUCKET_NAME).remove([storage_path])
    except Exception as e:
        st.warning(f"Impossibile eliminare il file dallo Storage: {e}")

# --- ANALISI LOCALE CON TESSERACT OCR & REGEX ---

def analyze_receipt_with_tesseract_ocr(file_bytes, mime_type, is_personal=False):
    try:
        if mime_type == "application/pdf":
            images = convert_from_bytes(file_bytes)
            if images:
                pil_img = images[0]
            else:
                return None
        else:
            pil_img = Image.open(io.BytesIO(file_bytes))

        pil_img = ImageOps.exif_transpose(pil_img).convert("L")
        text = pytesseract.image_to_string(pil_img, lang='ita')
        
        data_str = datetime.now().strftime("%Y-%m-%d")
        date_matches = re.findall(r'\b(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\b', text)
        if date_matches:
            d, m, y = date_matches[0]
            if len(y) == 2:
                y = "20" + y
            try:
                parsed_dt = datetime(int(y), int(m), int(d))
                data_str = parsed_dt.strftime("%Y-%m-%d")
            except ValueError:
                pass

        importo = 0.0
        lines = text.split('\n')
        total_candidates = []
        all_amounts = re.findall(r'\b\d+[\.,]\d{2}\b', text)
        if all_amounts:
            floats = [float(amt.replace(',', '.')) for amt in all_amounts]
            for line in lines:
                if any(k in line.lower() for k in ['totale', 'tot', 'eur', '€', 'importo', 'somma', 'euro']):
                    line_amounts = re.findall(r'\d+[\.,]\d{2}', line)
                    if line_amounts:
                        for la in line_amounts:
                            total_candidates.append(float(la.replace(',', '.')))
            if total_candidates:
                importo = max(total_candidates)
            else:
                importo = max(floats) if floats else 0.0

        destinazione = "Esercente Sconosciuto"
        valid_lines = [l.strip() for l in lines if len(l.strip()) > 3]
        for vl in valid_lines[:4]:
            if not re.match(r'^[\d\W]+$', vl) and not any(w in vl.lower() for w in ['ricevuta', 'scontrino', 'fattura', 'via', 'piazza', 'tel']):
                destinazione = vl
                break

        text_lower = text.lower()
        
        if is_personal:
            categoria_suggerita = "Spesa"
            pagamento_suggerito = "Card"
            if any(k in text_lower for k in ['deliveroo', 'just eat', 'glovo', 'pizza', 'sushi', 'ristorante']):
                categoria_suggerita = "Cibo/Delivery"
            elif any(k in text_lower for k in ['benzina', 'gasolio', 'eni', 'q8', 'autostrada']):
                categoria_suggerita = "Auto"
            elif any(k in text_lower for k in ['esselunga', 'conad', 'coop', 'lidl', 'supermercato']):
                categoria_suggerita = "Spesa"
        else:
            categoria_suggerita = "ALTRO"
            pagamento_suggerito = "Contanti"
            if any(k in text_lower for k in ['benzina', 'gasolio', 'q8', 'eni', 'agip']):
                categoria_suggerita = "CARBURANTE_CC"
                pagamento_suggerito = "Carta Carburante"
            elif any(k in text_lower for k in ['ristorante', 'pizzeria', 'bar', 'pranzo', 'cena']):
                categoria_suggerita = "RISTORANTI_CC"
                pagamento_suggerito = "CC (Carta)"

        return {
            "data": data_str,
            "destinazione": destinazione[:40],
            "importo": round(importo, 2),
            "categoria_suggerita": categoria_suggerita,
            "pagamento_suggerito": pagamento_suggerito,
            "note": text.replace('\n', ' ')[:120]
        }

    except Exception as e:
        st.error(f"❌ Errore OCR: {e}")
        return None

# --- GENERAZIONE DOCUMENTI (EXCEL & PDF) ---

def genera_excel(anno, modo, m_start, m_end):
    if not os.path.exists(EXCEL_TEMPLATE):
        st.error(f"File modello '{EXCEL_TEMPLATE}' non trovato!")
        return None

    wb = openpyxl.load_workbook(EXCEL_TEMPLATE)
    ROW_OFFSET = 5

    for m in range(m_start, m_end + 1):
        nome_foglio = MANDI_NOMI[m - 1]
        if nome_foglio not in wb.sheetnames:
            continue

        ws = wb[nome_foglio]
        start_date = f"{anno}-{m:02d}-01"
        end_date = f"{anno+1}-01-01" if m == 12 else f"{anno}-{m+1:02d}-01"

        response = supabase.table("spese") \
            .select("*") \
            .gte("data", start_date) \
            .lt("data", end_date) \
            .execute()
        
        spese = response.data

        for spesa in spese:
            d_str = spesa["data"]
            dest = spesa.get("destinazione")
            scopo = spesa.get("scopo")
            cat_key = spesa.get("categoria")
            imp = spesa.get("importo", 0.0)
            note = spesa.get("note")
            is_foreign = spesa.get("valuta_straniera", 0)

            dt = datetime.strptime(d_str, "%Y-%m-%d")
            giorno = dt.day
            target_row = giorno + ROW_OFFSET

            if dest:
                curr_dest = str(ws.cell(row=target_row, column=3).value or "").strip()
                if not curr_dest:
                    ws.cell(row=target_row, column=3, value=dest)
                else:
                    dest_list = [d.strip() for d in curr_dest.split('\\')]
                    if dest not in dest_list:
                        ws.cell(row=target_row, column=3, value=f"{curr_dest}\\{dest}")

            if scopo:
                curr_scopo = str(ws.cell(row=target_row, column=4).value or "").strip()
                if not curr_scopo:
                    ws.cell(row=target_row, column=4, value=scopo)
                else:
                    scopo_list = [s.strip() for s in curr_scopo.split('\\')]
                    if scopo not in scopo_list:
                        ws.cell(row=target_row, column=4, value=f"{curr_scopo}\\{scopo}")

            if note:
                curr_note = str(ws.cell(row=target_row, column=15).value or "").strip()
                if not curr_note:
                    ws.cell(row=target_row, column=15, value=note)
                else:
                    note_list = [n.strip() for n in curr_note.split(';')]
                    if note not in note_list:
                        ws.cell(row=target_row, column=15, value=f"{curr_note}; {note}")

            col_idx = CATEGORIA_COLONNA.get(cat_key)
            if col_idx and imp:
                imp_val = round(float(imp), 2)
                cell = ws.cell(row=target_row, column=col_idx)
                curr_val = cell.value

                if curr_val is None or curr_val == "":
                    cell.value = imp_val
                    if is_foreign == 1:
                        cell.fill = FILL_ORANGE
                elif isinstance(curr_val, (int, float)):
                    cell.value = f"={curr_val}+{imp_val}"
                    if is_foreign == 1 or cell.fill == FILL_ORANGE or cell.fill == FILL_YELLOW:
                        cell.fill = FILL_YELLOW
                elif isinstance(curr_val, str) and curr_val.startswith("="):
                    cell.value = f"{curr_val}+{imp_val}"
                    if is_foreign == 1 or cell.fill == FILL_ORANGE or cell.fill == FILL_YELLOW:
                        cell.fill = FILL_YELLOW

    if modo == 'anno':
        output_filename = f"Note_Spese_Anno_{anno}.xlsx"
    elif modo == 'range':
        output_filename = f"Note_Spese_{MANDI_NOMI[m_start-1]}_{MANDI_NOMI[m_end-1]}_{anno}.xlsx"
    else:
        output_filename = f"Note_Spese_{MANDI_NOMI[m_start-1]}_{anno}.xlsx"

    wb.save(output_filename)
    return output_filename

def genera_pdf_allegati(anno, mese):
    start_date = f"{anno}-{mese:02d}-01"
    end_date = f"{anno+1}-01-01" if mese == 12 else f"{anno}-{mese+1:02d}-01"

    response = supabase.table("spese") \
        .select("*") \
        .gte("data", start_date) \
        .lt("data", end_date) \
        .order("data", desc=False) \
        .order("id", desc=False) \
        .execute()
        
    spese = [s for s in response.data if s.get("allegato_path")]

    if not spese:
        return None

    pdf_filename = f"Allegati_Spese_{MANDI_NOMI[mese-1]}_{anno}.pdf"
    doc = SimpleDocTemplate(
        pdf_filename,
        pagesize=A4,
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=16, leading=20, alignment=1)
    card_body_style = ParagraphStyle('CardBody', parent=styles['Normal'], fontSize=8, leading=10, fontName="Helvetica")

    story = []
    story.append(Paragraph(f"<b>Allegati Spese (Grayscale) - {MANDI_NOMI[mese-1]} {anno}</b>", title_style))
    story.append(Spacer(1, 15))

    cells = []
    pdf_files_to_merge = []

    for spesa in spese:
        url = spesa["allegato_path"]
        is_pdf = url.lower().endswith(".pdf")

        d_fmt = datetime.strptime(spesa['data'], "%Y-%m-%d").strftime("%d/%m/%Y")
        dest = spesa.get('destinazione') or '-'
        cat = spesa.get('categoria') or '-'
        imp = f"€ {spesa.get('importo', 0.0):.2f}"
        note = spesa.get('note') or ''

        text_content = f"""
        <b>Data:</b> {d_fmt} | <b>Importo:</b> {imp}<br/>
        <b>Destinazione:</b> {dest}<br/>
        <b>Cat:</b> {cat}<br/>
        """
        if note:
            text_content += f"<b>Note:</b> {note}"

        try:
            resp = requests.get(url, timeout=10)
            if resp.status_code == 200:
                if is_pdf:
                    pdf_files_to_merge.append({
                        "bytes": resp.content,
                        "info": f"Allegato PDF - Data: {d_fmt} - Destinazione: {dest} - Importo: {imp}"
                    })
                    cell_elements = [
                        Paragraph(text_content, card_body_style),
                        Spacer(1, 4),
                        Paragraph("📄 <i>[Documento PDF allegato in coda al report]</i>", card_body_style)
                    ]
                    cells.append(cell_elements)
                else:
                    img_stream = io.BytesIO(resp.content)
                    pil_img = Image.open(img_stream)
                    pil_img = ImageOps.exif_transpose(pil_img)
                    pil_img = pil_img.convert("L")
                    
                    gray_stream = io.BytesIO()
                    pil_img.save(gray_stream, format="JPEG")
                    gray_stream.seek(0)

                    max_w, max_h = 240, 260
                    w, h = pil_img.size
                    ratio = min(max_w / w, max_h / h)
                    final_w, final_h = int(w * ratio), int(h * ratio)

                    rl_img = RLImage(gray_stream, width=final_w, height=final_h)

                    cell_elements = [
                        Paragraph(text_content, card_body_style),
                        Spacer(1, 4),
                        rl_img
                    ]
                    cells.append(cell_elements)

        except Exception:
            continue

    grid_data = []
    for i in range(0, len(cells), 2):
        row = [cells[i]]
        if i + 1 < len(cells):
            row.append(cells[i+1])
        else:
            row.append("")
        grid_data.append(row)

    if grid_data:
        t = Table(grid_data, colWidths=[275, 275])
        t.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
        ]))
        story.append(t)

    doc.build(story)

    if pdf_files_to_merge:
        merger = PdfWriter()
        merger.append(pdf_filename)

        for pdf_item in pdf_files_to_merge:
            pdf_bytes = io.BytesIO(pdf_item["bytes"])
            merger.append(pdf_bytes)

        merged_pdf_filename = f"Allegati_Spese_{MANDI_NOMI[mese-1]}_{anno}_Completo.pdf"
        merger.write(merged_pdf_filename)
        merger.close()
        
        if os.path.exists(pdf_filename):
            os.remove(pdf_filename)
        return merged_pdf_filename

    return pdf_filename

# --- INTERFACCIA STREAMLIT ---
st.title("🧾 Gestione Note Spese 2026")

tab1, tab2, tab3, tab4 = st.tabs(["➕ Nuova Spesa", "📑 Registri / Modifica", "📊 Report Spese", "📁 Export Excel & PDF"])

if "upload_key" not in st.session_state:
    st.session_state["upload_key"] = 0

if "ai_extracted_data" not in st.session_state:
    st.session_state["ai_extracted_data"] = {}

# --- TAB 1: REGISTRAZIONE SMART ---
with tab1:
    st.subheader("Registra Nuova Spesa")
    
    tipo_spesa = st.radio("Tipo di Spesa", ["💼 Lavoro", "🏡 Personale"], horizontal=True)
    is_personal = (tipo_spesa == "🏡 Personale")
    
    st.markdown("#### Step 1: Carica Allegato (Foto o PDF)")
    col_up1, col_up2 = st.columns(2)
    k_img = f"img_uploader_{st.session_state['upload_key']}"
    k_pdf = f"pdf_uploader_{st.session_state['upload_key']}"
    
    with col_up1:
        uploaded_image = st.file_uploader("📷 Scatta/Scegli Foto", type=["jpg", "png", "jpeg"], key=k_img)
    with col_up2:
        uploaded_pdf = st.file_uploader("📄 Scegli PDF", type=["pdf"], key=k_pdf)

    prepared_file = None

    if uploaded_image is not None:
        if uploaded_image.size > MAX_FILE_SIZE_MB * 1024 * 1024:
            st.error("⚠️ Il file immagine supera il limite di 5MB!")
        else:
            img = Image.open(uploaded_image)
            img = ImageOps.exif_transpose(img)
            enable_crop = st.checkbox("✂️ Attiva ritaglio/crop immagine", value=True)

            if enable_crop:
                cropped_img = st_cropper(img, realtime_update=True, box_color="#00FF00", aspect_ratio=None)
                cropped_bytes = optimize_pil_image(cropped_img)
            else:
                cropped_bytes = optimize_pil_image(img)
                st.image(cropped_bytes, caption="Foto scontrino intera", use_container_width=True)

            prepared_file = {"bytes": cropped_bytes, "ext": "jpg", "mime": "image/jpeg"}

    elif uploaded_pdf is not None:
        if uploaded_pdf.size > MAX_FILE_SIZE_MB * 1024 * 1024:
            st.error("⚠️ Il file PDF supera il limite di 5MB!")
        else:
            st.success(f"📄 PDF pronto: `{uploaded_pdf.name}`")
            prepared_file = {"bytes": uploaded_pdf.getvalue(), "ext": "pdf", "mime": "application/pdf"}

    if prepared_file:
        if st.button("🔍 Estrai Dati con OCR", type="secondary", use_container_width=True):
            with st.spinner("Estrazione testo e riconoscimento in corso..."):
                extracted = analyze_receipt_with_tesseract_ocr(prepared_file["bytes"], prepared_file["mime"], is_personal)
                if extracted:
                    st.session_state["ai_extracted_data"] = extracted
                    st.success("✅ Dati estratti con successo!")

    st.divider()
    st.markdown("#### Step 2: Verifica e Conferma Dati Spesa")

    ai_data = st.session_state.get("ai_extracted_data", {})
    default_date = datetime.now()
    if ai_data.get("data"):
        try:
            default_date = datetime.strptime(ai_data["data"], "%Y-%m-%d")
        except ValueError:
            pass

    col_d, col_w = st.columns([1, 1])
    with col_d:
        data_input = st.date_input("When (Data)", value=default_date)
    with col_w:
        dest_input = st.text_input("Where (Destinazione/Esercente)", value=ai_data.get("destinazione", ""))

    scopo_input = st.text_input("Why (Scopo)") if not is_personal else ""

    st.write("**What (Categoria)**")
    cat_options = CATEGORIE_PERSONALI if is_personal else CATEGORIE_LAVORO
    
    default_cat_idx = 0
    ai_cat = ai_data.get("categoria_suggerita", "")
    for idx, opt in enumerate(cat_options):
        if ai_cat.lower() in opt.lower():
            default_cat_idx = idx
            break

    cat_selection = st.selectbox("Categoria", options=cat_options, index=default_cat_idx) if is_personal else st.radio("Categoria", options=cat_options, index=default_cat_idx, horizontal=True, label_visibility="collapsed")

    is_telepass = not is_personal and ("Telepass" in cat_selection)

    if not is_telepass:
        st.write("**Payment (Metodo Pagamento)**")
        pay_options = PAGAMENTI_PERSONALI if is_personal else PAGAMENTI_LAVORO
        default_pay_idx = 0
        ai_pay = ai_data.get("pagamento_suggerito", "")
        for idx, opt in enumerate(pay_options):
            if ai_pay.lower() in opt.lower():
                default_pay_idx = idx
                break

        pay_selection = st.radio("Pagamento", options=pay_options, index=default_pay_idx, horizontal=True, label_visibility="collapsed")
    else:
        pay_selection = None
        st.info("ℹ️ Per il Telepass il metodo di pagamento viene omesso.")

    col_imp, col_curr = st.columns([2, 1])
    with col_imp:
        ai_imp = float(ai_data.get("importo", 0.0))
        importo_input = st.number_input("Value (€)", value=ai_imp, min_value=0.0, step=0.5, format="%.2f")
    with col_curr:
        st.write("")
        st.write("")
        is_foreign = st.checkbox("Non € (🔣)") if not is_personal else False

    note_input = st.text_area("Notes", value=ai_data.get("note", ""), height=70)

    if st.button("💾 Salva Spesa", type="primary", use_container_width=True):
        if importo_input <= 0 and not is_telepass:
            st.warning("Inserisci un importo valido!")
        else:
            d_str = data_input.strftime("%Y-%m-%d")
            
            file_url = None
            if not is_personal and prepared_file:
                file_url = upload_file_to_supabase(
                    file_bytes=prepared_file["bytes"],
                    data_spesa=d_str,
                    extension=prepared_file["ext"],
                    content_type=prepared_file["mime"]
                )

            if is_personal:
                new_record = {
                    "data": d_str,
                    "destinazione": dest_input,
                    "scopo": "",
                    "categoria": cat_selection,
                    "metodo_pagamento": pay_selection,
                    "importo": importo_input,
                    "note": note_input,
                    "valuta_straniera": 0
                }
                supabase.table("spese_personali").insert(new_record).execute()
            else:
                metodo_str = None if cat_selection == "TELEPASS" else pay_selection

                new_record = {
                    "data": d_str,
                    "destinazione": dest_input,
                    "scopo": scopo_input,
                    "categoria": cat_selection,
                    "metodo_pagamento": metodo_str,
                    "importo": importo_input,
                    "km": 0.0,
                    "note": note_input,
                    "allegato_path": file_url,
                    "valuta_straniera": 1 if is_foreign else 0
                }
                supabase.table("spese").insert(new_record).execute()
            
            st.success("✅ **Spesa registrata con successo nel database!**")
            st.session_state["ai_extracted_data"] = {}
            st.session_state["upload_key"] += 1
            time.sleep(1)
            st.rerun()

# --- TAB 2: CONSULTAZIONE & MODIFICA (RIPRISTINATO) ---
with tab2:
    st.subheader("Archivio Spese Registrate")
    tipo_archivio = st.radio("Seleziona archivio", ["Lavoro", "Personale"], horizontal=True, key="archivio_radio")
    
    table_name = "spese" if tipo_archivio == "Lavoro" else "spese_personali"
    
    try:
        response = supabase.table(table_name).select("*").order("data", desc=True).execute()
        data_list = response.data
    except Exception as e:
        st.error(f"❌ Errore di comunicazione con Supabase: {e}")
        data_list = []

    if not data_list:
        st.info("Nessuna spesa memorizzata in questo archivio.")
    else:
        df_spese = pd.DataFrame(data_list)
        
        st.write("💡 *Clicca su una riga della tabella per selezionarla e modificarla.*")
        
        event = st.dataframe(
            df_spese,
            use_container_width=True,
            on_select="rerun",
            selection_mode="single-row"
        )

        st.divider()
        st.subheader("Modifica / Elimina Spesa Selezionata")
        
        selected_rows = event.selection.rows if event and hasattr(event, 'selection') else []
        
        if selected_rows:
            selected_index = selected_rows[0]
            rec = df_spese.iloc[selected_index]
            record_id = rec['id']
            
            st.info(f"Stai modificando il record **ID #{record_id}** ({tipo_archivio}) del {rec['data']}")

            with st.form("edit_form"):
                e_data = st.date_input("Data", datetime.strptime(str(rec['data']), "%Y-%m-%d"))
                e_dest = st.text_input("Destinazione", rec.get('destinazione') or "")
                
                if tipo_archivio == "Lavoro":
                    e_scopo = st.text_input("Scopo", rec.get('scopo') or "")
                    e_cat = st.selectbox("Categoria", CATEGORIE_LAVORO, index=CATEGORIE_LAVORO.index(rec['categoria']) if rec['categoria'] in CATEGORIE_LAVORO else 0)
                    e_pay_list = ["- (Nessuno)"] + PAGAMENTI_LAVORO
                    curr_pay = rec.get('metodo_pagamento')
                    curr_idx = PAGAMENTI_LAVORO.index(curr_pay) + 1 if curr_pay in PAGAMENTI_LAVORO else 0
                    e_pay = st.selectbox("Pagamento", e_pay_list, index=curr_idx)
                    e_foreign = st.checkbox("Valuta non-€ (🔣)", value=bool(rec.get('valuta_straniera', 0)))
                else:
                    e_scopo = ""
                    e_cat = st.selectbox("Categoria", CATEGORIE_PERSONALI, index=CATEGORIE_PERSONALI.index(rec['categoria']) if rec['categoria'] in CATEGORIE_PERSONALI else 0)
                    e_pay_list = ["- (Nessuno)"] + PAGAMENTI_PERSONALI
                    curr_pay = rec.get('metodo_pagamento')
                    curr_idx = PAGAMENTI_PERSONALI.index(curr_pay) + 1 if curr_pay in PAGAMENTI_PERSONALI else 0
                    e_pay = st.selectbox("Pagamento", e_pay_list, index=curr_idx)
                    e_foreign = False

                e_imp = st.number_input("Importo (€)", value=float(rec['importo']), step=0.5)
                e_note = st.text_area("Note", rec.get('note') or "")

                if tipo_archivio == "Lavoro":
                    allegato = rec.get('allegato_path')
                    if isinstance(allegato, str) and allegato.startswith("http"):
                        if allegato.lower().endswith(".pdf"):
                            st.markdown(f"📄 [Visualizza Allegato PDF]({allegato})")
                        else:
                            st.image(allegato, caption="Allegato foto", width=250)
                
                c_sub, c_del = st.columns([1, 1])
                with c_sub:
                    save_mod = st.form_submit_button("💾 Salva Modifiche")
                with c_del:
                    delete_mod = st.form_submit_button("🗑️ Elimina Record", type="secondary")
                
                if save_mod:
                    if tipo_archivio == "Lavoro":
                        updated_record = {
                            "data": e_data.strftime("%Y-%m-%d"),
                            "destinazione": e_dest,
                            "scopo": e_scopo,
                            "categoria": e_cat,
                            "metodo_pagamento": None if e_pay == "- (Nessuno)" else e_pay,
                            "importo": e_imp,
                            "note": e_note,
                            "valuta_straniera": 1 if e_foreign else 0
                        }
                    else:
                        updated_record = {
                            "data": e_data.strftime("%Y-%m-%d"),
                            "destinazione": e_dest,
                            "categoria": e_cat,
                            "metodo_pagamento": None if e_pay == "- (Nessuno)" else e_pay,
                            "importo": e_imp,
                            "note": e_note
                        }
                    
                    supabase.table(table_name).update(updated_record).eq("id", int(record_id)).execute()
                    st.success("Record aggiornato con successo!")
                    time.sleep(1)
                    st.rerun()
                    
                if delete_mod:
                    if tipo_archivio == "Lavoro":
                        photo_url = rec.get('allegato_path')
                        if photo_url:
                            delete_photo_from_storage(photo_url)

                    supabase.table(table_name).delete().eq("id", int(record_id)).execute()
                    st.warning("Record eliminato con successo!")
                    time.sleep(1)
                    st.rerun()
        else:
            st.caption("👈 Seleziona una riga dalla tabella sopra per visualizzare il modulo di modifica/eliminazione.")

# --- TAB 3: REPORTISTICA & STATISTICHE PERSONALI ---
with tab3:
    st.subheader("📊 Report Spese & Statistiche")
    
    report_type = st.radio("Tipologia Report", ["Lavoro", "Personale"], horizontal=True, key="rep_type")
    
    if report_type == "Lavoro":
        col_rep_year, col_rep_month = st.columns([1, 1])
        with col_rep_year:
            rep_year = st.number_input("Anno", value=2026, step=1, key="rep_year_input")
        with col_rep_month:
            rep_month = st.selectbox("Mese da analizzare", range(1, 13), format_func=lambda x: MANDI_NOMI[x-1], index=datetime.now().month - 1, key="rep_m_input")
            
        start_d = f"{rep_year}-{rep_month:02d}-01"
        end_d = f"{rep_year+1}-01-01" if rep_month == 12 else f"{rep_year}-{rep_month+1:02d}-01"
        
        rep_res = supabase.table("spese").select("*").gte("data", start_d).lt("data", end_d).execute()
        rep_data = rep_res.data
        
        if not rep_data:
            st.info(f"Nessuna spesa trovata per **{MANDI_NOMI[rep_month-1]} {rep_year}**.")
        else:
            df_rep = pd.DataFrame(rep_data)
            totale_mese = df_rep['importo'].sum()
            spese_telepass = df_rep[df_rep['categoria'] == 'TELEPASS']['importo'].sum()
            spese_carb_carta = df_rep[df_rep['categoria'] == 'CARBURANTE_CARTA']['importo'].sum()
            
            totale_senza_telepass = totale_mese - spese_telepass
            totale_senza_telepass_e_carb = totale_mese - spese_telepass - spese_carb_carta

            st.markdown(f"### Totali per **{MANDI_NOMI[rep_month-1]} {rep_year}**")
            m1, m2 = st.columns(2)
            m1.metric("Totale Generale Mese", f"€ {totale_mese:.2f}")
            m2.metric("Totale (Senza Telepass)", f"€ {totale_senza_telepass:.2f}")
            
            m3, m4, m5 = st.columns(3)
            m3.metric("Totale (Senza Telepass & Carta Carb.)", f"€ {totale_senza_telepass_e_carb:.2f}")
            m4.metric("Totale Telepass", f"€ {spese_telepass:.2f}")
            m5.metric("Totale Carta Carburante", f"€ {spese_carb_carta:.2f}")
            
            st.divider()
            col_c1, col_c2 = st.columns(2)
            with col_c1:
                st.markdown("#### Per Categoria")
                cat_group = df_rep.groupby('categoria')['importo'].agg(['sum', 'count']).reset_index()
                cat_group.columns = ['Categoria', 'Totale (€)', 'N. Spese']
                st.dataframe(cat_group.sort_values(by='Totale (€)', ascending=False), hide_index=True, use_container_width=True)
            with col_c2:
                st.markdown("#### Per Metodo Pagamento")
                df_rep_pay = df_rep.copy()
                df_rep_pay['metodo_pagamento'] = df_rep_pay['metodo_pagamento'].fillna('- (Nessuno)')
                pay_group = df_rep_pay.groupby('metodo_pagamento')['importo'].agg(['sum', 'count']).reset_index()
                pay_group.columns = ['Metodo', 'Totale (€)', 'N. Spese']
                st.dataframe(pay_group.sort_values(by='Totale (€)', ascending=False), hide_index=True, use_container_width=True)
    else:
        st.markdown("### 🏡 Statistiche Spese Personali")
        res_pers = supabase.table("spese_personali").select("*").execute()
        if res_pers.data:
            df_p = pd.DataFrame(res_pers.data)
            df_p['data'] = pd.to_datetime(df_p['data'])
            
            oggi = datetime.now()
            mese_corrente = df_p[df_p['data'].dt.month == oggi.month]
            mese_precedente = df_p[df_p['data'].dt.month == (oggi.month - 1 if oggi.month > 1 else 12)]
            
            tot_mese_curr = mese_corrente['importo'].sum()
            tot_mese_prev = mese_precedente['importo'].sum()
            
            col_m1, col_m2 = st.columns(2)
            col_m1.metric("Spese Mese Corrente", f"€ {tot_mese_curr:.2f}", delta=f"€ {tot_mese_curr - tot_mese_prev:.2f} vs mese prec.")
            col_m2.metric("Spese Mese Precedente", f"€ {tot_mese_prev:.2f}")
            
            st.markdown("#### Spese per Categoria (Mese Corrente)")
            if not mese_corrente.empty:
                cat_sum = mese_corrente.groupby('categoria')['importo'].sum().reset_index()
                cat_sum.columns = ['Categoria', 'Totale (€)']
                st.dataframe(cat_sum.sort_values(by='Totale (€)', ascending=False), hide_index=True, use_container_width=True)
            else:
                st.info("Nessuna spesa personale registrata per il mese corrente.")
        else:
            st.info("Nessuna spesa personale registrata.")

# --- TAB 4: ESPORTAZIONE EXCEL & PDF ---
with tab4:
    st.subheader("📁 Generazione Documenti (Lavoro)")
    
    col_e1, col_e2 = st.columns(2)
    with col_e1:
        anno_exp = st.number_input("Anno Esportazione", value=2026, step=1, key="anno_exp_input")
    with col_e2:
        mese_exp_pdf = st.selectbox("Mese per PDF Allegati", range(1, 13), format_func=lambda x: MANDI_NOMI[x-1], index=datetime.now().month - 1, key="mese_exp_pdf_input")

    st.divider()

    col_btn1, col_btn2 = st.columns(2)

    with col_btn1:
        st.markdown("#### 📊 Modello Excel")
        tipo_exp = st.selectbox("Modalità Excel", ["Mese Singolo", "Range di Mesi", "Anno Completo"], key="tipo_exp_sel")
        
        if tipo_exp == "Mese Singolo":
            m_sel = st.selectbox("Seleziona Mese Excel", range(1, 13), format_func=lambda x: MANDI_NOMI[x-1], index=datetime.now().month - 1, key="m_sel_input")
            m_in, m_fi = m_sel, m_sel
            modo_str = 'singolo'
        elif tipo_exp == "Range di Mesi":
            col_m1, col_m2 = st.columns(2)
            with col_m1:
                m_in = st.selectbox("Da Mese", range(1, 13), format_func=lambda x: MANDI_NOMI[x-1], index=0, key="m_in_input")
            with col_m2:
                m_fi = st.selectbox("A Mese", range(1, 13), format_func=lambda x: MANDI_NOMI[x-1], index=datetime.now().month - 1, key="m_fi_input")
            modo_str = 'range'
        else:
            m_in, m_fi = 1, 12
            modo_str = 'anno'

        if st.button("📊 Genera Excel", type="primary", use_container_width=True):
            res_file = genera_excel(anno_exp, modo_str, m_in, m_fi)
            if res_file and os.path.exists(res_file):
                with open(res_file, "rb") as f:
                    st.download_button(
                        label="📥 Scarica Excel",
                        data=f,
                        file_name=res_file,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

    with col_btn2:
        st.markdown("#### 🖼️📄 PDF Allegati Cumulativo")
        st.caption(f"Unisce immagini e PDF del mese di **{MANDI_NOMI[mese_exp_pdf-1]} {anno_exp}**.")
        
        if st.button("🖼️ Genera PDF Completo", type="primary", use_container_width=True):
            with st.spinner("Generazione PDF allegati in corso..."):
                pdf_res = genera_pdf_allegati(anno_exp, mese_exp_pdf)
                
            if pdf_res and os.path.exists(pdf_res):
                with open(pdf_res, "rb") as f_pdf:
                    st.download_button(
                        label="📥 Scarica PDF Allegati",
                        data=f_pdf,
                        file_name=pdf_res,
                        mime="application/pdf",
                        use_container_width=True
                    )
            else:
                st.warning("Nessun allegato trovato per il mese selezionato.")
