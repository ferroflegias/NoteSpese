import os
import io
import re
from datetime import datetime
import requests
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from PIL import Image, ImageOps
import pytesseract
from pdf2image import convert_from_bytes
from supabase import create_client, Client

from fastapi import FastAPI, File, UploadFile, Form, HTTPException, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from pypdf import PdfWriter

# ReportLab per PDF
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

app = FastAPI(title="NoteSpese PWA API")

# Mappa cartella statici
app.mount("/static", StaticFiles(directory="static"), name="static")

SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY")
BUCKET_NAME = "allegati-spese"
EXCEL_TEMPLATE = "Note spese 2026.xlsx"

MANDI_NOMI = [
    "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
    "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"
]

CATEGORIA_COLONNA = {
    "TELEPASS": 6,               # Colonna F
    "NOLO": 7,                    # Colonna G
    "PARCHEGGI_CONTANTI": 8,     # Colonna H
    "PARCHEGGI_CC": 9,           # Colonna I
    "RISTORANTI_CONTANTI": 10,   # Colonna J
    "RISTORANTI_CC": 11,         # Colonna K
    "CARBURANTE_CC": 12,         # Colonna L
    "CARBURANTE_CARTA": 13,      # Colonna M
    "ALTRO": 14                  # Colonna N
}

FILL_ORANGE = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def get_supabase() -> Client:
    if not SUPABASE_URL or not SUPABASE_KEY:
        raise HTTPException(status_code=500, detail="Credenziali Supabase non configurate.")
    return create_client(SUPABASE_URL, SUPABASE_KEY)

# --- ROTTE PWA STATISCHE ---
@app.get("/")
async def serve_index():
    return FileResponse("static/index.html")

@app.get("/manifest.json")
async def serve_manifest():
    return FileResponse("static/manifest.json")

@app.get("/sw.js")
async def serve_sw():
    return FileResponse("static/sw.js", media_type="application/javascript")

# --- API CORE ---

@app.get("/api/spese")
async def get_spese():
    supabase = get_supabase()
    res = supabase.table("spese").select("*").order("data", desc=True).order("id", desc=True).execute()
    return res.data

@app.post("/api/ocr")
async def run_ocr(file: UploadFile = File(...)):
    contents = await file.read()
    mime_type = file.content_type
    try:
        if mime_type == "application/pdf":
            images = convert_from_bytes(contents)
            if not images:
                raise HTTPException(status_code=400, detail="Impossibile convertire il PDF.")
            pil_img = images[0]
        else:
            pil_img = Image.open(io.BytesIO(contents))

        pil_img = ImageOps.exif_transpose(pil_img).convert("L")
        text = pytesseract.image_to_string(pil_img, lang='ita')

        data_str = datetime.now().strftime("%Y-%m-%d")
        date_matches = re.findall(r'\b(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\b', text)
        if date_matches:
            d, m, y = date_matches[0]
            if len(y) == 2: y = "20" + y
            try: data_str = datetime(int(y), int(m), int(d)).strftime("%Y-%m-%d")
            except ValueError: pass

        all_amounts = re.findall(r'\b\d+[\.,]\d{2}\b', text)
        importo = max([float(amt.replace(',', '.')) for amt in all_amounts]) if all_amounts else 0.0

        destinazione = "Esercente Sconosciuto"
        lines = [l.strip() for l in text.split('\n') if len(l.strip()) > 3]
        for vl in lines[:4]:
            if not re.match(r'^[\d\W]+$', vl) and not any(w in vl.lower() for w in ['ricevuta', 'scontrino', 'fattura', 'via', 'piazza']):
                destinazione = vl
                break

        return {
            "data": data_str,
            "destinazione": destinazione[:40],
            "importo": round(importo, 2),
            "note": text.replace('\n', ' ')[:120]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Errore OCR: {str(e)}")

@app.post("/api/spese")
async def create_spesa(
    data: str = Form(...),
    destinazione: str = Form(""),
    scopo: str = Form(""),
    categoria: str = Form(...),
    metodo_pagamento: str = Form(None),
    importo: float = Form(...),
    note: str = Form(""),
    valuta_straniera: int = Form(0),
    file: UploadFile = File(None)
):
    supabase = get_supabase()
    file_url = None

    if file:
        file_bytes = await file.read()
        dt = datetime.strptime(data, "%Y-%m-%d")
        ext = "pdf" if file.content_type == "application/pdf" else "jpg"
        filename = f"rec_{data}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"
        storage_path = f"default_user/{dt.strftime('%Y-%m')}/{filename}"

        supabase.storage.from_(BUCKET_NAME).upload(
            path=storage_path,
            file=file_bytes,
            file_options={"content-type": file.content_type or "image/jpeg"}
        )
        file_url = supabase.storage.from_(BUCKET_NAME).get_public_url(storage_path)

    new_record = {
        "data": data,
        "destinazione": destinazione,
        "scopo": scopo,
        "categoria": categoria,
        "metodo_pagamento": metodo_pagamento,
        "importo": importo,
        "km": 0.0,
        "note": note,
        "allegato_path": file_url,
        "valuta_straniera": valuta_straniera
    }

    res = supabase.table("spese").insert(new_record).execute()
    return {"status": "success", "data": res.data}

@app.delete("/api/spese/{record_id}")
async def delete_spesa(record_id: int):
    supabase = get_supabase()
    rec = supabase.table("spese").select("allegato_path").eq("id", record_id).execute()
    if rec.data and rec.data[0].get("allegato_path"):
        url = rec.data[0]["allegato_path"]
        if BUCKET_NAME in url:
            storage_path = url.split(f"{BUCKET_NAME}/")[-1]
            supabase.storage.from_(BUCKET_NAME).remove([storage_path])

    supabase.table("spese").delete().eq("id", record_id).execute()
    return {"status": "deleted"}

@app.get("/api/export/excel")
async def export_excel(anno: int = 2026, modo: str = "singolo", m_start: int = 1, m_end: int = 1):
    if not os.path.exists(EXCEL_TEMPLATE):
        raise HTTPException(status_code=404, detail="File modello Excel non trovato.")

    wb = openpyxl.load_workbook(EXCEL_TEMPLATE)
    supabase = get_supabase()
    ROW_OFFSET = 5

    for m in range(m_start, m_end + 1):
        nome_foglio = MANDI_NOMI[m - 1]
        if nome_foglio not in wb.sheetnames: continue

        ws = wb[nome_foglio]
        start_date = f"{anno}-{m:02d}-01"
        end_date = f"{anno+1}-01-01" if m == 12 else f"{anno}-{m+1:02d}-01"

        spese = supabase.table("spese").select("*").gte("data", start_date).lt("data", end_date).order("data").order("id").execute().data or []

        for spesa in spese:
            dt = datetime.strptime(spesa["data"], "%Y-%m-%d")
            target_row = dt.day + ROW_OFFSET

            dest, scopo, cat_key, imp, note, is_foreign = (
                spesa.get("destinazione"), spesa.get("scopo"), spesa.get("categoria"),
                spesa.get("importo", 0.0), spesa.get("note"), spesa.get("valuta_straniera", 0)
            )

            if dest:
                curr = str(ws.cell(row=target_row, column=3).value or "").strip()
                ws.cell(row=target_row, column=3, value=dest if not curr else f"{curr}\\{dest}")

            col_idx = CATEGORIA_COLONNA.get(cat_key)
            if col_idx and imp:
                cell = ws.cell(row=target_row, column=col_idx)
                curr_val = cell.value
                imp_val = round(float(imp), 2)
                if not curr_val:
                    cell.value = imp_val
                    if is_foreign == 1: cell.fill = FILL_ORANGE
                else:
                    cell.value = f"={curr_val}+{imp_val}" if str(curr_val).startswith("=") else f"={curr_val}+{imp_val}"
                    cell.fill = FILL_YELLOW

    output_filename = f"Export_Spese_{anno}.xlsx"
    wb.save(output_filename)
    return FileResponse(output_filename, filename=output_filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/api/export/pdf")
async def export_pdf(anno: int = 2026, mese: int = 1):
    supabase = get_supabase()
    start_date = f"{anno}-{mese:02d}-01"
    end_date = f"{anno+1}-01-01" if mese == 12 else f"{anno}-{mese+1:02d}-01"

    spese = supabase.table("spese").select("*").gte("data", start_date).lt("data", end_date).order("data").order("id").execute().data or []
    spese_con_allegato = [s for s in spese if s.get("allegato_path")]

    if not spese_con_allegato:
        raise HTTPException(status_code=404, detail="Nessun allegato trovato per il mese selezionato.")

    pdf_filename = f"Allegati_Spese_{MANDI_NOMI[mese-1]}_{anno}.pdf"
    doc = SimpleDocTemplate(pdf_filename, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=16, alignment=1)
    card_body_style = ParagraphStyle('CardBody', parent=styles['Normal'], fontSize=8, leading=10)

    story = [Paragraph(f"<b>Allegati Spese ({len(spese_con_allegato)} Voci) - {MANDI_NOMI[mese-1]} {anno}</b>", title_style), Spacer(1, 15)]
    cells, pdf_files_to_merge = [], []

    for spesa in spese_con_allegato:
        url = str(spesa["allegato_path"]).strip()
        is_pdf = url.lower().split("?")[0].endswith(".pdf")
        rec_id, d_fmt = spesa.get('id', '-'), datetime.strptime(spesa['data'], "%Y-%m-%d").strftime("%d/%m/%Y")
        text_content = f"<b>[ID #{rec_id}] Data:</b> {d_fmt} | <b>Importo:</b> € {spesa.get('importo', 0.0):.2f}<br/><b>Destinazione:</b> {spesa.get('destinazione') or '-'}<br/>"

        try:
            resp = requests.get(url, timeout=20)
            if resp.status_code == 200:
                if is_pdf:
                    pdf_files_to_merge.append({"bytes": resp.content})
                    cells.append([Paragraph(text_content, card_body_style), Spacer(1, 4), Paragraph("📄 <i>[Documento PDF allegato in coda]</i>", card_body_style)])
                else:
                    pil_img = Image.open(io.BytesIO(resp.content))
                    pil_img = ImageOps.exif_transpose(pil_img).convert("L")
                    gray_stream = io.BytesIO()
                    pil_img.save(gray_stream, format="JPEG")
                    gray_stream.seek(0)
                    cells.append([Paragraph(text_content, card_body_style), Spacer(1, 4), RLImage(gray_stream, width=200, height=220)])
        except Exception:
            continue

    grid_data = [[cells[i], cells[i+1]] if i+1 < len(cells) else [cells[i], ""] for i in range(0, len(cells), 2)]
    if grid_data:
        t = Table(grid_data, colWidths=[275, 275])
        t.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP'), ('ALIGN', (0,0), (-1,-1), 'CENTER')]))
        story.append(t)

    doc.build(story)

    if pdf_files_to_merge:
        merger = PdfWriter()
        merger.append(pdf_filename)
        for pdf_item in pdf_files_to_merge:
            merger.append(io.BytesIO(pdf_item["bytes"]))
        merged_pdf = f"Allegati_{MANDI_NOMI[mese-1]}_{anno}_Completo.pdf"
        with open(merged_pdf, "wb") as f_out: merger.write(f_out)
        merger.close()
        return FileResponse(merged_pdf, filename=merged_pdf, media_type="application/pdf")

    return FileResponse(pdf_filename, filename=pdf_filename, media_type="application/pdf")